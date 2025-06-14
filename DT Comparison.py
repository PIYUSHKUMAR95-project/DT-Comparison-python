import dash
from dash import dcc, html, Input, Output, State, ctx, dash_table
import pandas as pd
import base64
import io
import os
import tempfile
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

app = dash.Dash(__name__)
app.title = "Excel File Comparator"

temp_dir = tempfile.gettempdir()
file1_path = os.path.join(temp_dir, "uploaded_file1.xlsx")
file2_path = os.path.join(temp_dir, "uploaded_file2.xlsx")

# UI
app.layout = html.Div([
    html.H2("Excel File Comparator", style={"textAlign": "center"}),

    html.Div([
        html.Div([
            html.Label("Upload File 1 (.xls, .xlsx, .csv):"),
            dcc.Upload(
                id='upload-file1',
                children=html.Div(['📂 Drag & Drop or Click to Upload File 1']),
                multiple=False,
                style={
                    'width': '100%', 'padding': '10px',
                    'border': '2px dashed #999', 'borderRadius': '5px',
                    'textAlign': 'center', 'marginBottom': '20px'
                },
            ),
        ], style={"width": "45%", "display": "inline-block", "paddingRight": "5%"}),

        html.Div([
            html.Label("Upload File 2 (.xls, .xlsx, .csv):"),
            dcc.Upload(
                id='upload-file2',
                children=html.Div(['📂 Drag & Drop or Click to Upload File 2']),
                multiple=False,
                style={
                    'width': '100%', 'padding': '10px',
                    'border': '2px dashed #999', 'borderRadius': '5px',
                    'textAlign': 'center'
                },
            ),
        ], style={"width": "45%", "display": "inline-block"}),
    ]),

    html.Button("Compare Files", id="compare-button", n_clicks=0, style={"marginTop": "20px"}),
    html.Div(id='error-message', style={"color": "red", "marginTop": "10px"}),

    html.Hr(),
    html.Div(id='comparison-table')
])

def save_uploaded_file(contents, filename, temp_name):
    content_type, content_string = contents.split(',')
    decoded = base64.b64decode(content_string)
    extension = filename.split('.')[-1].lower()

    path = os.path.join(tempfile.gettempdir(), temp_name + ".xlsx")

    if extension == 'csv':
        df = pd.read_csv(io.StringIO(decoded.decode('utf-8')))
        df.to_excel(path, index=False)
    elif extension in ['xls', 'xlsx']:
        with open(path, 'wb') as f:
            f.write(decoded)
    else:
        return None
    return path

# Callback
@app.callback(
    Output('comparison-table', 'children'),
    Output('error-message', 'children'),
    Input('compare-button', 'n_clicks'),
    State('upload-file1', 'contents'),
    State('upload-file1', 'filename'),
    State('upload-file2', 'contents'),
    State('upload-file2', 'filename')
)
def compare_files(n_clicks, contents1, filename1, contents2, filename2):
    if not n_clicks:
        return dash.no_update, ""

    if not contents1 or not contents2:
        return None, "❌ Error: Please upload both files."

    try:
        path1 = save_uploaded_file(contents1, filename1, "uploaded_file1")
        path2 = save_uploaded_file(contents2, filename2, "uploaded_file2")

        if not path1 or not path2:
            return None, "❌ Error: Unsupported file format."

        # ========== START: YOUR COMPARISON LOGIC ==========
        sheet1 = pd.ExcelFile(path1).sheet_names[0]
        sheet2 = pd.ExcelFile(path2).sheet_names[0]

        df1 = pd.read_excel(path1, sheet_name=sheet1)
        df2 = pd.read_excel(path2, sheet_name=sheet2)

        column_mapping = {
            'IA_Code': ['IA_Code_1', 'IA_CODE1_DESC_NEW'],
            'Quantity': ['QUANTITY', 'FINAL_LETTERSHOP_QTY'],
            'PRIMARY_SOURCE_CODE': ['PRIMARY_SOURCE_CODE', 'PRIMARY_SOURCE_CODE'],
            'PRIMARY_SPID': ['PRIMARY_SPID', 'PRIMARY_SPID1_NEW'],
            'CAMPAIGN_CODE': ['CAMPAIGN_CODE', 'CAMPAIGN_CODE'],
            'TEMPLATE_CODE': ['TEMPLATE_CODE', 'TEMPLATE_CODE'],
            'EXPIRATION_DATE': ['EXPIRATION_DATE', 'EXPIRATION_DATE'],
            'PRESCREEN_DATE': ['PRESCREEN_DATE', 'PRESCREEN_DATE'],
            'POID': ['POID', 'POID'],
            'CELL_ID': ['CELL_ID', 'CELL_ID']
        }

        file1_rename = {}
        file2_rename = {}
        for std_name, (f1_col, f2_col) in column_mapping.items():
            if f1_col in df1.columns:
                file1_rename[f1_col] = std_name
            if f2_col in df2.columns:
                file2_rename[f2_col] = std_name

        df1.rename(columns=file1_rename, inplace=True)
        df2.rename(columns=file2_rename, inplace=True)

        common_columns = list(set(file1_rename.values()) & set(file2_rename.values()))
        if not common_columns:
            common_columns = df1.columns.intersection(df2.columns).tolist()

        if 'CELL_ID' in df1.columns:
            df1['CELL_ID'] = pd.to_numeric(df1['CELL_ID'], errors='ignore')
            df1 = df1.sort_values(by='CELL_ID').reset_index(drop=True)
        if 'CELL_ID' in df2.columns:
            df2['CELL_ID'] = pd.to_numeric(df2['CELL_ID'], errors='ignore')
            df2 = df2.sort_values(by='CELL_ID').reset_index(drop=True)

        def normalize_date(val):
            if pd.isnull(val):
                return None
            val_str = str(val).strip()
            if val_str.isdigit() and 7 <= len(val_str) <= 8:
                try:
                    padded = val_str.zfill(8)
                    return datetime.strptime(padded, "%m%d%Y").date()
                except:
                    pass
            for fmt in ("%m/%d/%Y", "%Y%m%d", "%m%d%Y"):
                try:
                    return datetime.strptime(val_str, fmt).date()
                except ValueError:
                    continue
            return val

        for col in common_columns:
            if "date" in col.lower():
                df1[col] = df1[col].apply(normalize_date)
                df2[col] = df2[col].apply(normalize_date)

        min_len = min(len(df1), len(df2))
        df1 = df1.iloc[:min_len].reset_index(drop=True)
        df2 = df2.iloc[:min_len].reset_index(drop=True)
        result_df = df1.copy()

        if 'Quantity' in df1.columns and 'Quantity' in df2.columns:
            diff_series = pd.to_numeric(df1['Quantity'], errors='coerce') - pd.to_numeric(df2['Quantity'], errors='coerce')
            insert_loc = result_df.columns.get_loc('Quantity') + 1
            result_df.insert(insert_loc, 'Quantity_Diff', diff_series)

        for col in common_columns:
            col1_vals = df1[col].astype(str).str.strip()
            col2_vals = df2[col].astype(str).str.strip()
            result_df[col] = [
                v1 if v1 == v2 else f'DIFF: {v1} | {v2}'
                for v1, v2 in zip(col1_vals, col2_vals)
            ]

        # Save results to file1 path
        book = load_workbook(path1)
        with pd.ExcelWriter(path1, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            writer._book = book
            result_df.to_excel(writer, sheet_name='Comparison_Result', index=False)

        wb = load_workbook(path1)
        ws = wb['Comparison_Result']
        red_font = Font(color="FF0000")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('DIFF:'):
                    cell.font = red_font

        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2
        wb.save(path1)
        # ========== END: COMPARISON LOGIC ==========

        # Filter only mismatches for UI
        mismatch_df = result_df[result_df.apply(lambda row: any(str(v).startswith("DIFF:") for v in row), axis=1)]
        if mismatch_df.empty:
            return html.Div("✅ No mismatches found."), ""

        return dash_table.DataTable(
            columns=[{"name": i, "id": i} for i in mismatch_df.columns],
            data=mismatch_df.to_dict('records'),
            style_table={'overflowX': 'auto', 'marginTop': '20px'},
            style_cell={'textAlign': 'left', 'fontFamily': 'Arial', 'padding': '5px'},
            style_data_conditional=[
                {
                    'if': {'filter_query': '{' + col + '}.contains("DIFF:")', 'column_id': col},
                    'color': 'red'
                } for col in mismatch_df.columns
            ]
        ), ""

    except Exception as e:
        return None, f"❌ Error: {str(e)}"

if __name__ == '__main__':
    app.run(debug=True)
