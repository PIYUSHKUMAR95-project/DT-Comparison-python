import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime

def compare_excels(file1_path, file2_path):
    try:
        # Load first sheet from both files
        sheet1 = pd.ExcelFile(file1_path).sheet_names[0]
        sheet2 = pd.ExcelFile(file2_path).sheet_names[0]
        print(f"Sheet 1: {sheet1}, Sheet 2: {sheet2}")

        # Read data
        df1 = pd.read_excel(file1_path, sheet_name=sheet1)
        df2 = pd.read_excel(file2_path, sheet_name=sheet2)
        print(f"File 1 shape: {df1.shape}, File 2 shape: {df2.shape}")

        # Optional column mapping (adjust as needed)
        column_mapping = {
            'IA_Code': ['IA_Code_1', 'IA_CODE1_DESC_NEW'],
            'Quantity': ['QUANTITY', 'FINAL_LETTERSHOP_QTY'],
            'PRIMARY_SOURCE_CODE': ['PRIMARY_SOURCE_CODE', 'PRIMARY_SOURCE_CODE'],
            'PRIMARY_SPID': ['PRIMARY_SPID', 'PRIMARY_SPID1_NEW'],
            'CAMPAIGN_CODE': ['CAMPAIGN_CODE', 'CAMPAIGN_CODE'],
            'TEMPLATE_CODE': ['TEMPLATE_CODE', 'TEMPLATE_CODE'],
            'EXPIRATION_DATE': ['EXPIRATION_DATE', 'EXPIRATION_DATE'],
            'PRESCREEN_DATE': ['PRESCREEN_DATE', 'PRESCREEN_DATE'],
            'POID': ['POID', 'POID'],
            'CELL_ID': ['CELL_ID', 'CELL_ID']
        }

        # Rename columns based on mapping
        file1_rename = {}
        file2_rename = {}
        for std_name, (f1_col, f2_col) in column_mapping.items():
            if f1_col in df1.columns:
                file1_rename[f1_col] = std_name
            if f2_col in df2.columns:
                file2_rename[f2_col] = std_name
        df1.rename(columns=file1_rename, inplace=True)
        df2.rename(columns=file2_rename, inplace=True)
        print(":pushpin: File1 Columns After Renaming:", df1.columns.tolist())
        print(":pushpin: File2 Columns After Renaming:", df2.columns.tolist())

        # Detect common columns
        common_columns = list(set(file1_rename.values()) & set(file2_rename.values()))
        if not common_columns:
            print("\n:warning: No common columns found from mapping. Trying auto-matching based on identical names.")
            common_columns = df1.columns.intersection(df2.columns).tolist()
        print(":white_check_mark: Common columns used for comparison:", common_columns)

        if not common_columns:
            raise ValueError("No common columns found for comparison.")

        # Sort both DataFrames by CELL_ID (handle numeric or string)
        if 'CELL_ID' in df1.columns:
            try:
                df1['CELL_ID'] = pd.to_numeric(df1['CELL_ID'], errors='ignore')
            except:
                pass
            df1 = df1.sort_values(by='CELL_ID').reset_index(drop=True)
        if 'CELL_ID' in df2.columns:
            try:
                df2['CELL_ID'] = pd.to_numeric(df2['CELL_ID'], errors='ignore')
            except:
                pass
            df2 = df2.sort_values(by='CELL_ID').reset_index(drop=True)

        # Normalize date fields if necessary
        def normalize_date(val):
            if pd.isnull(val):
                return None
            val_str = str(val).strip()
            if val_str.isdigit() and 7 <= len(val_str) <= 8:
                try:
                    padded = val_str.zfill(8)
                    return datetime.strptime(padded, "%m%d%Y").date()
                except:
                    pass
            for fmt in ("%m/%d/%Y", "%Y%m%d", "%m%d%Y"):
                try:
                    return datetime.strptime(val_str, fmt).date()
                except ValueError:
                    continue
            return val

        for col in common_columns:
            if "date" in col.lower():
                df1[col] = df1[col].apply(normalize_date)
                df2[col] = df2[col].apply(normalize_date)

        # Align row lengths
        min_len = min(len(df1), len(df2))
        df1 = df1.iloc[:min_len].reset_index(drop=True)
        df2 = df2.iloc[:min_len].reset_index(drop=True)
        # Create result DataFrame with all File 1 data
        result_df = df1.copy()

        # Adding a new column next to Quantity as Quantity Diff for the Quantity difference of file1 and file2
        if 'Quantity' in df1.columns and 'Quantity' in df2.columns:
            diff_series = pd.to_numeric(df1['Quantity'], errors='coerce') - pd.to_numeric(df2['Quantity'], errors='coerce')
            insert_loc = result_df.columns.get_loc('Quantity') + 1
            result_df.insert(insert_loc, 'Quantity_Diff', diff_series)

        # Compare and flag differences
        for col in common_columns:
            col1_vals = df1[col].astype(str)
            col2_vals = df2[col].astype(str)
            result_df[col] = [
                v1 if v1 == v2 else f'DIFF: {v1} | {v2}'
                for v1, v2 in zip(col1_vals, col2_vals)
            ]

        print(f":large_green_circle: Compared columns: {common_columns}")
        print(f":white_check_mark: Final result will contain all File 1 columns: {result_df.columns.tolist()}")

        # Write to Excel
        book = load_workbook(file1_path)
        with pd.ExcelWriter(file1_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            writer._book = book
            result_df.to_excel(writer, sheet_name='Comparison_Result', index=False)

        # Format output (color differences red)
        wb = load_workbook(file1_path)
        ws = wb['Comparison_Result']
        red_font = Font(color="FF0000")
        for row in ws.iter_rows(min_row=2):  # Skip header
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('DIFF:'):
                    cell.font = red_font

        # Auto-adjust column width
        for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        # Save workbook
        wb.save(file1_path)
        print("\n:white_check_mark: Comparison completed and saved in 'Comparison_Result' sheet.")
        return file1_path
    except Exception as e:
        print(f"Error in compare_excels: {str(e)}")
        raise
result_path = compare_excels('YourFile1.xlsx', 'YourFile2.xlsx')
print('Compared Excel saved at:', result_path)     
